from typing import Dict, List, Tuple, Type

import logging

logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)

console_handler = logging.StreamHandler()
file_handler = logging.FileHandler('eudoxa.log', 'w', 'utf-8')

console_handler.setLevel(logging.DEBUG)
file_handler.setLevel(logging.DEBUG)

formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
console_handler.setFormatter(formatter)
file_handler.setFormatter(formatter)

logger.addHandler(console_handler)
logger.addHandler(file_handler)


PROJ = "|PROJ|"
ASP = "|ASP| "
CONS = "|CONS|"
DOM = "|DOM|"
VDCM = "|VDCM|"

TRUE = "⊒"
FALSE = "⋣"
UNDEFINED = ""

BT = "≻"
BTE = "⪰"
EQ = "∼"
WTE = "⪯"
WT = "≺"

AL_RELATION_OPTIONS = [UNDEFINED, BT, BTE, EQ, WTE, WT]

DELTA = "Δ"
ZDIFF_TUPLE   = (None, None)  # internal key for natural zero-diffs
ZDIFF_DISPLAY = "◬"   # U+25EC — used in app views and Excel labels

GT = "⊐"
GTE = "⊒"
DEQ = "≜"
LTE = "⊑"
LT = "⊏"

VDIFF_RELATION_OPTIONS = [UNDEFINED, GT, GTE, DEQ, LTE, LT]

class Aspect:
    def __init__(self, name: str, data_type: Type, description: str = None):
        self.name = name
        self.data_type = data_type
        self.description = description
        self.levels: Dict[str, str] = {}
        self.vdiffs: List[VDiff] = [VDiff(name, None, None)]
        
    def add_level(self, level: str, description: str):
        logger.info(f"Adding level '{level}' to aspect '{self.name}'")
        try:
            parse_type(level, self.data_type)
        except (ValueError, TypeError) as e:
            raise ValueError(
                f"Level '{level}' is not valid for aspect '{self.name}' "
                f"with data_type '{self.data_type.__name__}': {e}"
            )
        for l_key in self.levels.keys():
            vd = VDiff(self.name, l_key, level)
            self.vdiffs.append(vd)
            vd_inv = VDiff(self.name, level, l_key)
            self.vdiffs.append(vd_inv)
        self.levels[level] = description

    def change_type(self, new_type: Type) -> list:
        """Change data_type if all existing levels are valid for new_type.
        Returns list of level names that fail to parse (empty means success)."""
        failing = []
        for level in self.levels:
            try:
                parse_type(level, new_type)
            except (ValueError, TypeError):
                failing.append(level)
        if not failing:
            self.data_type = new_type
        return failing

    def add_description(self, description: str):
        self.description = description

    def set_level_description(self, level: str, description: str):
        if level not in self.levels:
            raise ValueError(f"Level '{level}' does not exist in aspect '{self.name}'.")
        self.levels[level] = description

    def __repr__(self):
        return (f"Aspect(name='{self.name}', data_type='{self.data_type.__name__}', "
                f"description='{self.description}', levels={list(self.levels.keys())})")
    
    def to_dict(self):
        return {
            "name": self.name,
            "data_type": self.data_type.__name__,
            "description": self.description,
            "levels": self.levels,  # dict[str, str]
            "vdiffs": [
                {
                    "from_level": vd.from_level,
                    "to_level": vd.to_level
                }
                for vd in self.vdiffs
            ]
        }

    @classmethod
    def from_dict(cls, data):
        asp = cls(
            name=data["name"],
            data_type=str_to_type(data["data_type"]),
            description=data.get("description")
        )
    
        # Restore levels
        asp.levels = dict(data.get("levels", {}))
    
        # Restore vdiffs (replaces default)
        asp.vdiffs = [
            VDiff(
                aspect_name=data["name"],
                from_level=vd["from_level"],
                to_level=vd["to_level"]
            )
            for vd in data.get("vdiffs", [])
        ]
    
        return asp

class Consequence:
    def __init__(self, aspect_levels=None):
        self.aspect_levels: Dict[str, str] = aspect_levels if aspect_levels is not None else {}

    def __eq__(self, other):
        if not isinstance(other, Consequence):
            return NotImplemented
        return self.aspect_levels == other.aspect_levels

    def __getitem__(self, aspect_name: str) -> str:
        return self.aspect_levels.get(aspect_name, None)
    
    def __setitem__(self, aspect_name: str, level: str):
        self.aspect_levels[aspect_name] = level
    
    def __repr__(self):
        return "⟨" + ", ".join(f"{v}" for v in self.aspect_levels.values()) + "⟩"
    
    def to_dict(self):
        """Serialize Consequence to a plain dict."""
        return {
            "aspect_levels": self.aspect_levels
        }
    
    @classmethod
    def from_dict(cls, data):
        """Create a Consequence object from a dict with key 'aspect_levels'."""
        # Guard against None or missing key — fall back to empty dict
        aspect_levels = data.get("aspect_levels", {}) if isinstance(data, dict) else {}
        # Ensure all values are str or None
        cleaned = {}
        for k, v in aspect_levels.items():
            cleaned[str(k)] = None if v is None else str(v)
        return cls(cleaned)

class VDiff:
    
    def __init__(self, aspect_name: str, from_level: str, to_level: str):
        self.aspect_name = aspect_name
        self.from_level = from_level
        self.to_level = to_level
        
    def __eq__(self, other):
        if not isinstance(other, VDiff):
            return NotImplemented
        return (self.aspect_name == other.aspect_name and
                self.from_level == other.from_level and
                self.to_level == other.to_level)
    
    def __ne__(self, other):
        return not self.__eq__(other)

    def __hash__(self):
        return hash((self.aspect_name, self.from_level, self.to_level))

    def inv(self):
        return VDiff(self.aspect_name, self.to_level, self.from_level)

    def equals(self, other):
        if self.natural_zero() and other.natural_zero():
            return True
        return (self == other)

    def natural_zero(self):
        return self.from_level == self.to_level

    def __repr__(self):
        if self.natural_zero():
            return ZDIFF_DISPLAY
        return f"{DELTA}({self.from_level},{self.to_level})"

# The single canonical natural-zero vdiff used as a dict key in the vdcm.
# All aspect-specific natural-zero VDiff objects (where from_level == to_level)
# are normalised to this sentinel before any vdcm lookup or write.
NATURAL_ZERO = VDiff(None, None, None)

def str_to_type(data_type_str: str) -> Type:
    if data_type_str == 'int':
        return int
    elif data_type_str == 'float':
        return float
    return str

def parse_type(data_str: str, data_type: Type):
    if data_type == str:
        return str(data_str)
    elif data_type == int:
        return int(data_str)
    elif data_type == float:
        return float(data_str)
    return str(data_str)

def _vdiff_key(vd: VDiff) -> VDiff:
    """Normalise any natural-zero vdiff to the single NATURAL_ZERO sentinel."""
    return NATURAL_ZERO if vd.natural_zero() else vd

def get_vdiff_relation(vdcm, vd1: VDiff, vd2: VDiff,
                       default: str = UNDEFINED) -> str:
    """Look up the relation between vd1 and vd2 in vdcm.
    Natural-zero vdiffs are normalised to NATURAL_ZERO before lookup.
    Returns *default* (UNDEFINED) when the pair is absent."""
    row = vdcm.get(_vdiff_key(vd1))
    if row is None:
        return default
    return row.get(_vdiff_key(vd2), default)

def set_vdiff_relation(vdcm, vd1: VDiff, vd2: VDiff, new_rel: str) -> Tuple:
    """Write a relation into vdcm[vd1][vd2], normalising natural zeros.
    Both entries must already exist (initialised by expand_vdiff_comparison_matrix).
    Returns (add, coll) — one of which is always None."""
    k1, k2 = _vdiff_key(vd1), _vdiff_key(vd2)
    old_rel = vdcm[k1][k2]          # KeyError → not initialised (programming error)
    add, coll = None, None
    if new_rel == old_rel:           # Same relation — no change
        pass
    elif new_rel == UNDEFINED:       # Unset (clear) relation
        add = [vd1, new_rel, vd2]
        vdcm[k1][k2] = new_rel
    elif old_rel == UNDEFINED:       # Set new relation
        add = [vd1, new_rel, vd2]
        vdcm[k1][k2] = new_rel
    else:                            # Collision: TRUE ↔ FALSE
        coll = [vd1, old_rel, vd2, new_rel]
    return (add, coll)

def pos(vd: VDiff, aspect: Aspect, vdcm) -> bool:
    """vd ⊐ ◬  (strictly positive: vd ⊒ ◬ and ◬ ⋣ vd)."""
    if vd.natural_zero():
        return False
    return (get_vdiff_relation(vdcm, vd, NATURAL_ZERO) == TRUE and
            get_vdiff_relation(vdcm, NATURAL_ZERO, vd) == FALSE)

def non_neg(vd: VDiff, aspect: Aspect, vdcm) -> bool:
    """vd ⊒ ◬  (non-negative)."""
    if vd.natural_zero():
        return True   # ◬ ⊒ ◬ by definition
    return get_vdiff_relation(vdcm, vd, NATURAL_ZERO) == TRUE

def zero(vd: VDiff, aspect: Aspect, vdcm) -> bool:
    """vd ≜ ◬  (equivalent to zero: vd ⊒ ◬ and ◬ ⊒ vd)."""
    if vd.natural_zero():
        return True
    return (get_vdiff_relation(vdcm, vd, NATURAL_ZERO) == TRUE and
            get_vdiff_relation(vdcm, NATURAL_ZERO, vd) == TRUE)

def non_pos(vd: VDiff, aspect: Aspect, vdcm) -> bool:
    """◬ ⊒ vd  (non-positive)."""
    if vd.natural_zero():
        return True
    return get_vdiff_relation(vdcm, NATURAL_ZERO, vd) == TRUE

def neg(vd: VDiff, aspect: Aspect, vdcm) -> bool:
    """vd ⋣ ◬  (negative)."""
    if vd.natural_zero():
        return False  # ◬ is never negative
    return get_vdiff_relation(vdcm, vd, NATURAL_ZERO) == FALSE

def classify_vdiffs(asp: Aspect, vdcm) -> dict:
    """
    Classify all VDiffs for the given aspect into three mutually exclusive buckets:
      'non_negative': Δ(X,Y) ⊒ ◬
      'negative':     Δ(X,Y) ⋣ ◬
      'undecided':    no relation to ◬ set
    Returns a dict with those three keys, each mapping to a list of VDiff objects
    in asp.vdiffs order (natural zero-diff first, then by level insertion order).
    """
    result = {"non_negative": [], "negative": [], "undecided": []}
    for vd in asp.vdiffs:
        if non_neg(vd, asp, vdcm):
            result["non_negative"].append(vd)
        elif neg(vd, asp, vdcm):
            result["negative"].append(vd)
        else:
            result["undecided"].append(vd)
    return result

def app_ac(origin: List, ac: Tuple, adds: List, colls: List) -> Tuple:
    (add, coll) = ac
    if add:
        adds.append([origin[0], origin[1], add])
    if coll:
        colls.append([origin[0], origin[1], coll])
    return (adds, colls)

from itertools import product

import openpyxl

class EudoxaManager:
    
    def __init__(self):
        logger.info('Initializing EudoxaManager')
        self.aspects: Dict[str, Aspect] = {}
        self.consequences : Dict[str, Consequence]  = {}
        self.vdiff_comparison_matrix: Dict[VDiff, Dict[VDiff, str]] = {}

    def has_aspect(self, aspect_name: str) -> bool:
        return aspect_name in self.aspects

    def add_aspect(self, name: str, data_type_str: str, description = None) -> Aspect:
        logger.info(f"Adding aspect '{name}'")
        if '|' in str(name):
            raise ValueError(f"Aspect name '{name}' may not contain '|'.")
        if name in self.aspects:
            raise ValueError(f"Aspect '{name}' already exists.")
        data_type = str_to_type(data_type_str)
        self.aspects[name] = Aspect(name, data_type, description)
        # Mark all existing consequences as incomplete for the new aspect
        for consequence in self.consequences.values():
            if name not in consequence.aspect_levels:
                consequence.aspect_levels[name] = None
        return self.aspects[name]

    def get_aspect(self, name: str) -> Aspect:
        return self.aspects[name]

    def create_aspect_level_relations_graph(self, aspect_name: str,
                                             use_closure: bool = True,
                                             use_tr: bool = True):
        """Build a DAG of aspect level relations.
        Nodes are equivalence classes (sets of mutually EQ levels).
        Edges run from better classes to worse classes.

        use_closure=True  — compute aspect-local transitive closure (Warshall)
                            before building the graph, so transitively derived
                            relations are included. No full vdiff closure needed.
        use_closure=False — use only the explicitly set matrix entries.
        use_tr=True       — apply transitive reduction before returning.
        use_tr=False      — return the full graph without reduction.
        """
        import networkx as nx
        aspect = self.get_aspect(aspect_name)
        levels = list(aspect.levels.keys())
        if not levels:
            return nx.DiGraph()

        # Read the aspect-local relation matrix
        rel = {
            (la, lb): self.get_aspect_level_relation(aspect_name, la, lb)
            for la in levels for lb in levels
        }

        if use_closure:
            # Composition table: compose(r1, r2) -> inferred relation or None
            def compose(r1, r2):
                if r1 == EQ:  return r2
                if r2 == EQ:  return r1
                if r1 == BT  and r2 in (BT, BTE): return BT
                if r1 == BTE and r2 == BTE:        return BTE
                if r1 == WT  and r2 in (WT, WTE):  return WT
                if r1 == WTE and r2 == WTE:         return WTE
                return None

            # Strength ordering for picking the stronger of two relations
            _strength = {BT: 5, WT: 5, BTE: 3, WTE: 3, EQ: 1, UNDEFINED: 0}
            def stronger(r1, r2):
                return r1 if _strength.get(r1, 0) >= _strength.get(r2, 0) else r2

            # Warshall's algorithm on the aspect-local matrix
            for pivot in levels:
                for a in levels:
                    for c in levels:
                        if a == c:
                            continue
                        inferred = compose(rel[(a, pivot)], rel[(pivot, c)])
                        if inferred:
                            rel[(a, c)] = stronger(rel[(a, c)], inferred)

        def get_rel(la, lb):
            return rel[(la, lb)]

        # Step 1: compute equivalence classes (X ~ Y iff EQ)
        eq_classes = nx.equivalence_classes(
            levels,
            lambda x, y: get_rel(x, y) == EQ
        )

        # Assign a stable label and pick a representative for each class
        class_list = [sorted(cls) for cls in eq_classes]
        class_list.sort()
        rep   = {tuple(cls): cls[0] for cls in class_list}
        label = {tuple(cls): f" {EQ} ".join(cls) for cls in class_list}

        # Step 2: build graph on equivalence classes
        nxdg = nx.DiGraph()
        for cls in class_list:
            nxdg.add_node(tuple(cls))
        for cls1 in class_list:
            for cls2 in class_list:
                if cls1 == cls2:
                    continue
                r1, r2 = rep[tuple(cls1)], rep[tuple(cls2)]
                if get_rel(r1, r2) in (BT, BTE):
                    nxdg.add_edge(tuple(cls1), tuple(cls2))

        # Step 3: transitive reduction (optional)
        if use_tr:
            nxdg = nx.transitive_reduction(nxdg)

        # Attach label and members as node attributes for callers
        for cls in class_list:
            key = tuple(cls)
            if key in nxdg.nodes:
                nxdg.nodes[key]['label']   = label[key]
                nxdg.nodes[key]['members'] = cls
        return nxdg

    def show_aspect_level_relations_graph(self, aspect_name, nxdg):
        import networkx as nx
        import matplotlib as mpl
        import matplotlib.pyplot as plt
        for layer, nodes in enumerate(nx.topological_generations(nxdg)):
            for node in nodes:
                nxdg.nodes[node]["layer"] = layer
        pos = nx.multipartite_layout(nxdg, subset_key="layer", align="horizontal")
        fig, ax = plt.subplots()
        nx.draw_networkx(nxdg, pos=pos, ax=ax, node_size = 800)
        ax.set_title("Aspect level relations graph for " + aspect_name)
        plt.show()
        
    def add_consequence(self, short_name: str, aspect_levels: Dict[str,str]) -> List:
        added_aspect_levels = []
        if short_name in self.consequences:
            raise ValueError(f"A consequence named '{short_name}' already exists.")

        if not aspect_levels.keys() == self.aspects.keys():
            raise ValueError("Aspect keys do not match.")

        # Normalise to strings for comparison (mirrors what is stored)
        normalised = {k: str(v) for k, v in aspect_levels.items()}
        for existing_name, existing_cons in self.consequences.items():
            if all(existing_cons[a] == normalised[a] for a in self.aspects):
                raise ValueError(
                    f"Consequence '{short_name}' is identical to "
                    f"existing consequence '{existing_name}'."
                )

        c = Consequence(aspect_levels)
        logger.debug(f"Adding '{short_name}' to consequence set.")
        for aspect_name, level in aspect_levels.items():
            aspect = self.get_aspect(aspect_name)
            level_str = str(level)
            if level_str not in aspect.levels:
                self.add_aspect_level(aspect_name, level_str, None)
                added_aspect_levels.append(aspect_name + ":" + level_str)
            c.__setitem__(aspect_name, level_str)

        self.consequences[short_name] = c
        return added_aspect_levels

    def remove_consequence(self, short_name: str):
        if short_name in self.consequences:
            del self.consequences[short_name]

    def set_consequence_level(self, short_name: str, aspect_name: str, level: str) -> None:
        """Set the level for one aspect in a named consequence.
        Only accepts levels that already exist in the aspect.
        Raises ValueError if the update would create a duplicate consequence.
        """
        if short_name not in self.consequences:
            raise ValueError(f"Consequence '{short_name}' not found.")
        if aspect_name not in self.aspects:
            raise ValueError(f"Aspect '{aspect_name}' not found.")
        level_str = str(level)
        if level_str not in self.aspects[aspect_name].levels:
            raise ValueError(f"Level '{level_str}' not found in aspect '{aspect_name}'.")
        # Build the updated aspect_levels dict and check uniqueness
        updated = dict(self.consequences[short_name].aspect_levels)
        updated[aspect_name] = level_str
        for other_name, other_cons in self.consequences.items():
            if other_name == short_name:
                continue
            if all(other_cons.aspect_levels.get(a) == updated.get(a) for a in self.aspects):
                raise ValueError(
                    f"Setting this level would make '{short_name}' identical to '{other_name}'."
                )
        self.consequences[short_name].aspect_levels[aspect_name] = level_str

    @property
    def incomplete_consequences(self) -> dict:
        """Return {short_name: [aspect_names_with_None]} for incomplete consequences."""
        result = {}
        for name, cons in self.consequences.items():
            missing = [asp for asp in self.aspects if cons.aspect_levels.get(asp) is None]
            if missing:
                result[name] = missing
        return result

    def dom(self, ca: Consequence, cb: Consequence) -> bool:
        # TODO: Error handling
        n_bt = 0
        n_bte = 0
        for an in self.aspects:
            la = ca.__getitem__(an)
            lb = cb.__getitem__(an)
            rel = self.get_aspect_level_relation(an, la, lb)
            if rel == WT:
                return False
            if rel == WTE:
                pass
            elif rel == EQ:
                n_bte += 1
            elif rel == BTE:
                n_bte += 1
            elif rel == BT:
                n_bt += 1
            else: # Undefined
                return NotImplemented
        if n_bt + n_bte == len(self.aspects) and n_bt>0:
            return True
        return None

    def dom_possible(self, ca: Consequence, cb: Consequence,
                     aspect_rel: dict) -> str:
        """Three-valued dominance check using pre-built aspect-local matrices.

        Returns:
          'confirmed' — all relations defined, ca dominates cb
          'possible'  — no confirmed worse relation (WT/WTE) and at least one
                        UNDEFINED relation (dominance can't be ruled out)
          'none'      — confirmed non-dominance (WT/WTE found), or all
                        relations defined but no dominance
        """
        n_bt  = 0
        n_bte = 0
        has_undefined = False
        for an in self.aspects:
            la  = ca[an]
            lb  = cb[an]
            rel = aspect_rel[an][(la, lb)]
            if rel in (WT, WTE):
                return 'none'
            elif rel == UNDEFINED:
                has_undefined = True
            elif rel == EQ:
                n_bte += 1
            elif rel == BTE:
                n_bte += 1
            elif rel == BT:
                n_bt += 1
        if has_undefined:
            return 'possible'
        # All defined — standard dominance check
        if n_bt + n_bte == len(self.aspects) and n_bt > 0:
            return 'confirmed'
        return 'none'

    def show_dominance_graph(self, nxdg, html_file: str):
        import networkx as nx
        import matplotlib as mpl
        import matplotlib.pyplot as plt
        for layer, nodes in enumerate(nx.topological_generations(nxdg)):
            for node in nodes:
                nxdg.nodes[node]["layer"] = layer
        pos = nx.multipartite_layout(nxdg, subset_key="layer", align="horizontal")
        fig, ax = plt.subplots()
        nx.draw_networkx(nxdg, pos=pos, ax=ax, node_size = 4000)
        ax.set_title("Consequence dominance graph")
        plt.show()

    def get_computable_consequences(self, use_closure: bool = False):
        """Partition named consequences into those for which all pairwise
        aspect level relations are defined, and those that are excluded.

        Returns (included, excluded) where:
          included  — list of short_name strings
          excluded  — dict {short_name: [reason_string, ...]}

        A consequence ca is excluded if there exists any other consequence cb
        and any aspect for which rel(ca[aspect], cb[aspect]) is UNDEFINED.

        When use_closure=True, the aspect-local Warshall closure is used
        to check relations, so transitively derived relations count.
        """
        names  = list(self.consequences.keys())
        if not names:
            return [], {}

        # Build aspect-local closures (or raw matrices) once per aspect
        aspect_rel = {}   # {asp_name: {(la, lb): rel}}
        for asp_name, aspect in self.aspects.items():
            levels = list(aspect.levels.keys())
            rel = {
                (la, lb): self.get_aspect_level_relation(asp_name, la, lb)
                for la in levels for lb in levels
            }
            if use_closure:
                def compose(r1, r2):
                    if r1 == EQ:  return r2
                    if r2 == EQ:  return r1
                    if r1 == BT  and r2 in (BT, BTE):  return BT
                    if r1 == BTE and r2 == BTE:         return BTE
                    if r1 == WT  and r2 in (WT, WTE):   return WT
                    if r1 == WTE and r2 == WTE:          return WTE
                    return None
                _strength = {BT: 5, WT: 5, BTE: 3, WTE: 3, EQ: 1, UNDEFINED: 0}
                def stronger(r1, r2):
                    return r1 if _strength.get(r1, 0) >= _strength.get(r2, 0) else r2
                for pivot in levels:
                    for a in levels:
                        for c in levels:
                            if a == c: continue
                            inferred = compose(rel[(a, pivot)], rel[(pivot, c)])
                            if inferred:
                                rel[(a, c)] = stronger(rel[(a, c)], inferred)
            aspect_rel[asp_name] = rel

        # Find excluded consequences: those with any UNDEFINED relation to another
        excluded = {}   # {short_name: set of reason strings}
        for na, ca in self.consequences.items():
            for nb, cb in self.consequences.items():
                if na == nb:
                    continue
                for asp_name in self.aspects:
                    la = ca[asp_name]
                    lb = cb[asp_name]
                    if aspect_rel[asp_name][(la, lb)] == UNDEFINED:
                        reason = f"{asp_name}: {la} vs {lb} (from {nb})"
                        excluded.setdefault(na, set()).add(reason)

        # Convert reason sets to sorted lists
        excluded = {k: sorted(v) for k, v in excluded.items()}
        included = [n for n in names if n not in excluded]
        return included, excluded

    def create_dominance_graph(self, use_tr: bool = True):
        """Build a dominance graph with confirmed and possible edges.

        Returns a dict:
          nodes             — list of {id, complete} where complete=True means
                              all pairwise relations to all other nodes are defined
          edges_confirmed   — edges where dominance is confirmed (transitive-
                              reduced if use_tr=True and graph is a DAG)
          edges_possible    — edges where dominance is possible (never reduced)
        """
        import networkx as nx

        consequences = self.consequences
        if not consequences:
            return {"nodes": [], "edges_confirmed": [], "edges_possible": []}

        # Build aspect-local raw relation matrices once
        aspect_rel = {}
        for asp_name, aspect in self.aspects.items():
            levels = list(aspect.levels.keys())
            aspect_rel[asp_name] = {
                (la, lb): self.get_aspect_level_relation(asp_name, la, lb)
                for la in levels for lb in levels
            }

        # Classify each node as complete or incomplete
        names = list(consequences.keys())
        node_complete = {}
        for na in names:
            ca = consequences[na]
            complete = all(
                aspect_rel[an][(ca[an], consequences[nb][an])] != UNDEFINED
                for nb in names if nb != na
                for an in self.aspects
            )
            node_complete[na] = complete

        # Classify each directed pair
        confirmed_pairs = []
        possible_pairs  = []
        for na in names:
            for nb in names:
                if na == nb:
                    continue
                result = self.dom_possible(
                    consequences[na], consequences[nb], aspect_rel
                )
                if result == 'confirmed':
                    confirmed_pairs.append((str(consequences[na]),
                                            str(consequences[nb])))
                elif result == 'possible':
                    possible_pairs.append((str(consequences[na]),
                                           str(consequences[nb])))

        # Optionally apply transitive reduction to confirmed edges
        str_to_name = {str(c): n for n, c in consequences.items()}
        if use_tr and confirmed_pairs:
            g = nx.DiGraph()
            for n in names:
                g.add_node(str(consequences[n]))
            g.add_edges_from(confirmed_pairs)
            if nx.is_directed_acyclic_graph(g):
                g = nx.transitive_reduction(g)
                confirmed_pairs = list(g.edges)
            # If not a DAG (shouldn't happen with consistent data), keep all

        nodes = [
            {"id": str(consequences[n]), "name": n,
             "complete": node_complete[n]}
            for n in names
        ]
        return {
            "nodes":           nodes,
            "edges_confirmed": confirmed_pairs,
            "edges_possible":  possible_pairs
        }

    def create_dominance_table(self) -> Dict[Tuple[str, str], bool]:
        dom_table = {}
        for ca in self.consequences.values():
            for cb in self.consequences.values():
                dom_tf = self.dom(ca, cb)
                if dom_tf:
                    cacb = (str(ca), str(cb))
                    dom_table[cacb] = dom_tf
        return dom_table

    def export_dominance_table_to_excel(self,
                                        dom_table: Dict[Tuple[str, str], bool],
                                        filename: str):
        import openpyxl
        title = DOM
        try:
            wb = openpyxl.load_workbook(filename)
        except: # Create new workbook if file does not exist
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = title
        try:
            ws = wb[title]
        except: # Create new sheet if it does not exist
            ws = wb.create_sheet(title=title)
        self.export_dominance_table_to_worksheet(dom_table, ws)
        wb.save(filename)
        
    def export_dominance_table_to_worksheet(self, dom_table: Dict[Tuple[str, str], bool], ws):
        # Row 1: headers
        ws["A1"] = 'From Type'
        ws["B1"] = 'From Name'
        ws["C1"] = 'Edge Type'
        ws["D1"] = 'To Type'
        ws["E1"] = 'To Name'
        row_index = 2
        for ((c1, c2), v) in dom_table.items():
            ws.cell(row=row_index, column=1).value='Consequence'
            ws.cell(row=row_index, column=2).value=c1
            ws.cell(row=row_index, column=3).value='DOM'
            ws.cell(row=row_index, column=4).value='Consequence'
            ws.cell(row=row_index, column=5).value=c2
            row_index += 1

    def set_level_description(self, aspect_name: str, level_name: str, description: str):
        logger.debug(f"Setting description for level '{level_name}' in aspect '{aspect_name}'.")
        if aspect_name not in self.aspects:
            raise ValueError(f"Aspect '{aspect_name}' does not exist.")
        self.aspects[aspect_name].set_level_description(level_name, description)

    def add_aspect_level(self, aspect_name: str, level, description: str):
        logger.debug(f"Adding level '{level}' to aspect '{aspect_name}'.")
        if aspect_name not in self.aspects:
            raise ValueError(f"Aspect '{aspect_name}' does not exist.")
        aspect = self.get_aspect(aspect_name)
        level_str = str(level)
        if level_str in aspect.levels:
            raise ValueError(f"Level '{level_str}' already exists in aspect '{aspect_name}'.")
        # Add the new level
        aspect.add_level(level_str, description)
        # Update the value-difference comparison matrix
        self.expand_vdiff_comparison_matrix(aspect_name)

    def stage_remove_aspect_level(self, aspect_name: str, level: str) -> dict:
        """Compute the impact of removing a level without committing.

        Returns a dict:
          vdiffs_removed        — repr strings of VDiffs that will disappear
          al_relations_unset    — [{la, relation, lb}] for set within-aspect relations
          vdcm_entries_removed  — [{vd1, relation, vd2}] for non-UNDEFINED cross-aspect
                                   VDCM entries (excludes NATURAL_ZERO-backed AL relations)
          consequences_removed  — short names of consequences that will be deleted
        """
        if aspect_name not in self.aspects:
            raise ValueError(f"Aspect '{aspect_name}' does not exist.")
        aspect = self.aspects[aspect_name]
        if level not in aspect.levels:
            raise ValueError(f"Level '{level}' does not exist in aspect '{aspect_name}'.")

        vdiffs_to_remove = [vd for vd in aspect.vdiffs
                            if vd.from_level == level or vd.to_level == level]
        vdiff_keys = {_vdiff_key(vd) for vd in vdiffs_to_remove}

        # Within-aspect AL relations being unset
        al_relations = []
        for other in aspect.levels:
            if other == level:
                continue
            rel = self.get_aspect_level_relation(aspect_name, level, other)
            if rel not in (UNDEFINED,) and rel is not NotImplemented:
                al_relations.append({"la": level, "relation": rel, "lb": other})

        # Cross-aspect VDCM entries being removed (exclude NATURAL_ZERO and other
        # deleted VDiffs since those are already covered by al_relations_unset)
        vdcm_entries = []
        vdcm = self.vdiff_comparison_matrix
        seen = set()
        for k1 in vdiff_keys:
            row = vdcm.get(k1, {})
            for k2, rel in row.items():
                if rel == UNDEFINED or k2 == NATURAL_ZERO or k2 in vdiff_keys:
                    continue
                pair = (repr(k1), repr(k2))
                if pair not in seen:
                    seen.add(pair)
                    vdcm_entries.append({"vd1": repr(k1), "relation": rel, "vd2": repr(k2)})
        for k_other, row in vdcm.items():
            if k_other in vdiff_keys or k_other == NATURAL_ZERO:
                continue
            for k1 in vdiff_keys:
                rel = row.get(k1, UNDEFINED)
                if rel == UNDEFINED:
                    continue
                pair = (repr(k_other), repr(k1))
                if pair not in seen:
                    seen.add(pair)
                    vdcm_entries.append({"vd1": repr(k_other), "relation": rel, "vd2": repr(k1)})

        consequences_removed = [
            name for name, cons in self.consequences.items()
            if cons[aspect_name] == level
        ]

        return {
            "vdiffs_removed":       [repr(vd) for vd in vdiffs_to_remove],
            "al_relations_unset":   al_relations,
            "vdcm_entries_removed": vdcm_entries,
            "consequences_removed": consequences_removed,
        }

    def confirm_remove_aspect_level(self, aspect_name: str, level: str):
        """Remove an aspect level and all associated VDCM entries and consequences."""
        if aspect_name not in self.aspects:
            raise ValueError(f"Aspect '{aspect_name}' does not exist.")
        aspect = self.aspects[aspect_name]
        if level not in aspect.levels:
            raise ValueError(f"Level '{level}' does not exist in aspect '{aspect_name}'.")

        vdiffs_to_remove = [vd for vd in aspect.vdiffs
                            if vd.from_level == level or vd.to_level == level]
        vdiff_keys = {_vdiff_key(vd) for vd in vdiffs_to_remove}

        # Remove VDCM rows and columns for the deleted VDiffs
        vdcm = self.vdiff_comparison_matrix
        for k in list(vdiff_keys):
            vdcm.pop(k, None)
        for row in vdcm.values():
            for k in list(vdiff_keys):
                row.pop(k, None)

        # Remove level from aspect
        del aspect.levels[level]
        aspect.vdiffs = [vd for vd in aspect.vdiffs
                         if vd.from_level != level and vd.to_level != level]

        # Remove consequences that use this level
        to_delete = [name for name, cons in self.consequences.items()
                     if cons[aspect_name] == level]
        for name in to_delete:
            del self.consequences[name]

    def stage_remove_aspect(self, aspect_name: str) -> dict:
        """Compute the impact of removing an entire aspect without committing.

        Returns a dict:
          levels_removed        — level names
          vdiffs_removed        — repr strings of all VDiffs for this aspect
          al_relations_unset    — [{la, relation, lb}] for set within-aspect relations
          vdcm_entries_removed  — [{vd1, relation, vd2}] for non-UNDEFINED cross-aspect
                                   VDCM entries involving this aspect's VDiffs
          consequence_count     — total number of named consequences
          duplicate_groups      — [{keep, discard}] groups where removal causes duplicates
          discarded_if_keep     — number of consequences dropped under "keep" option
        """
        if aspect_name not in self.aspects:
            raise ValueError(f"Aspect '{aspect_name}' does not exist.")
        aspect = self.aspects[aspect_name]

        levels_removed = list(aspect.levels.keys())
        vdiff_keys = {_vdiff_key(vd) for vd in aspect.vdiffs
                      if _vdiff_key(vd) != NATURAL_ZERO}

        # Within-aspect AL relations that are set
        al_relations = []
        seen_al = set()
        for la in aspect.levels:
            for lb in aspect.levels:
                if la == lb:
                    continue
                rel = self.get_aspect_level_relation(aspect_name, la, lb)
                if rel not in (UNDEFINED,) and rel is not NotImplemented:
                    pair = (la, lb)
                    if pair not in seen_al:
                        seen_al.add(pair)
                        al_relations.append({"la": la, "relation": rel, "lb": lb})

        # Non-UNDEFINED cross-aspect VDCM entries involving this aspect's VDiffs
        vdcm = self.vdiff_comparison_matrix
        vdcm_entries = []
        seen_vdcm = set()
        for k1 in vdiff_keys:
            for k2, rel in vdcm.get(k1, {}).items():
                if rel == UNDEFINED or k2 in vdiff_keys:
                    continue
                pair = (repr(k1), repr(k2))
                if pair not in seen_vdcm:
                    seen_vdcm.add(pair)
                    vdcm_entries.append({"vd1": repr(k1), "relation": rel, "vd2": repr(k2)})
        for k_other, row in vdcm.items():
            if k_other in vdiff_keys:
                continue
            for k1 in vdiff_keys:
                rel = row.get(k1, UNDEFINED)
                if rel == UNDEFINED:
                    continue
                pair = (repr(k_other), repr(k1))
                if pair not in seen_vdcm:
                    seen_vdcm.add(pair)
                    vdcm_entries.append({"vd1": repr(k_other), "relation": rel, "vd2": repr(k1)})

        # Duplicate groups after removing the aspect key from all consequences
        from collections import defaultdict
        groups: dict = defaultdict(list)
        remaining_aspects = [a for a in self.aspects if a != aspect_name]
        for short_name, cons in self.consequences.items():
            key = tuple(cons.aspect_levels.get(a) for a in remaining_aspects)
            groups[key].append(short_name)

        duplicate_groups = []
        for names in groups.values():
            if len(names) > 1:
                names_sorted = sorted(names)
                duplicate_groups.append({
                    "keep":    names_sorted[0],
                    "discard": names_sorted[1:],
                })

        discarded_if_keep = sum(len(g["discard"]) for g in duplicate_groups)

        return {
            "levels_removed":        levels_removed,
            "vdiffs_removed":        [repr(vd) for vd in aspect.vdiffs
                                      if _vdiff_key(vd) != NATURAL_ZERO],
            "al_relations_unset":    al_relations,
            "vdcm_entries_removed":  vdcm_entries,
            "consequence_count":     len(self.consequences),
            "duplicate_groups":      duplicate_groups,
            "discarded_if_keep":     discarded_if_keep,
        }

    def confirm_remove_aspect(self, aspect_name: str, consequences: str) -> None:
        """Remove an aspect and all associated data.

        consequences — one of:
          "keep"              strip aspect key, keep one per duplicate group
                              (lexicographically first short name)
          "discard_duplicates" strip aspect key, discard every member of any
                              duplicate group, keep only unique consequences
          "discard_all"       delete all named consequences
        """
        if aspect_name not in self.aspects:
            raise ValueError(f"Aspect '{aspect_name}' does not exist.")
        aspect = self.aspects[aspect_name]
        vdiff_keys = {_vdiff_key(vd) for vd in aspect.vdiffs
                      if _vdiff_key(vd) != NATURAL_ZERO}

        # Remove VDCM rows and columns for this aspect's VDiffs
        vdcm = self.vdiff_comparison_matrix
        for k in list(vdiff_keys):
            vdcm.pop(k, None)
        for row in vdcm.values():
            for k in list(vdiff_keys):
                row.pop(k, None)

        # Remove aspect from model
        del self.aspects[aspect_name]

        if consequences == "discard_all":
            self.consequences.clear()
            return

        # Strip aspect key from every consequence
        for cons in self.consequences.values():
            cons.aspect_levels.pop(aspect_name, None)

        # Group by remaining tuple to find duplicates
        from collections import defaultdict
        remaining_aspects = list(self.aspects.keys())
        groups: dict = defaultdict(list)
        for short_name, cons in self.consequences.items():
            key = tuple(cons.aspect_levels.get(a) for a in remaining_aspects)
            groups[key].append(short_name)

        to_delete = set()
        for names in groups.values():
            if len(names) > 1:
                if consequences == "discard_duplicates":
                    # Discard every member of the duplicate group
                    to_delete.update(names)
                else:  # "keep"
                    # Discard all but the lexicographically first
                    for name in sorted(names)[1:]:
                        to_delete.add(name)

        for name in to_delete:
            del self.consequences[name]

    def set_aspect_level_relation(self, aspect: str, la, lb, rel: str) -> Tuple:
        adds, colls = [], []
        a = self.get_aspect(aspect)
        a_type = a.data_type
        la_str, lb_str = str(la), str(lb)
        if a is None:
            raise ValueError(f"Aspect '{aspect}' does not exist.")
        if not la_str in a.levels:
            raise ValueError(f"Aspect level '{la}' [{a_type}] does not exist.")
        if not lb_str in a.levels:
            raise ValueError(f"Aspect level '{lb}' [{a_type}] does not exist.")
        zero = NATURAL_ZERO
        vd_ab = VDiff(aspect, la_str, lb_str)
        vd_ba = VDiff(aspect, lb_str, la_str)

        origin = ['SETREL', [aspect, la_str, rel, lb_str]]
        if rel == UNDEFINED:
            app_ac(origin, self.set_vdiff_relation(vd_ab, zero, UNDEFINED), adds, colls)
            app_ac(origin, self.set_vdiff_relation(vd_ba, zero, UNDEFINED), adds, colls)
            app_ac(origin, self.set_vdiff_relation(zero, vd_ab, UNDEFINED), adds, colls)
            app_ac(origin, self.set_vdiff_relation(zero, vd_ba, UNDEFINED), adds, colls)
        elif rel == BT:
            app_ac(origin, self.set_vdiff_relation(vd_ab, zero, TRUE), adds, colls)
            app_ac(origin, self.set_vdiff_relation(vd_ba, zero, FALSE), adds, colls)
            app_ac(origin, self.set_vdiff_relation(zero, vd_ab, FALSE), adds, colls)
            app_ac(origin, self.set_vdiff_relation(zero, vd_ba, TRUE), adds, colls)
        elif rel == BTE:
            app_ac(origin, self.set_vdiff_relation(vd_ab, zero, TRUE), adds, colls)
            app_ac(origin, self.set_vdiff_relation(zero, vd_ba, TRUE), adds, colls)
        elif rel == EQ:
            app_ac(origin, self.set_vdiff_relation(vd_ab, zero, TRUE), adds, colls)
            app_ac(origin, self.set_vdiff_relation(vd_ba, zero, TRUE), adds, colls)
            app_ac(origin, self.set_vdiff_relation(zero, vd_ab, TRUE), adds, colls)
            app_ac(origin, self.set_vdiff_relation(zero, vd_ba, TRUE), adds, colls)
        elif rel == WTE:
            app_ac(origin, self.set_vdiff_relation(vd_ba, zero, TRUE), adds, colls)
            app_ac(origin, self.set_vdiff_relation(zero, vd_ab, TRUE), adds, colls)
        elif rel == WT:
            app_ac(origin, self.set_vdiff_relation(vd_ba, zero, TRUE), adds, colls)
            app_ac(origin, self.set_vdiff_relation(vd_ab, zero, FALSE), adds, colls)
            app_ac(origin, self.set_vdiff_relation(zero, vd_ba, FALSE), adds, colls)
            app_ac(origin, self.set_vdiff_relation(zero, vd_ab, TRUE), adds, colls)
        return (adds, colls)

    def try_set_aspect_level_relation(self, aspect: str, la, lb, rel: str) -> Tuple:
        """Validate and commit a relation addition using the closure as a staging area.

        Flow:
          1. Compute the current closure from the matrix.
          2. Apply the requested addition to the closure (not the matrix).
          3. If that causes an immediate collision, reject and return it.
          4. Recompute the closure of the staged dict.
          5. If the recomputed closure has collisions, reject and return them.
          6. If clean, commit only the explicit addition to the matrix.
             Return direct adds and any inferred additions from the closure.
        """
        import copy
        a = self.get_aspect(aspect)
        a_type = a.data_type
        la_str, lb_str = str(la), str(lb)
        if a is None:
            raise ValueError(f"Aspect '{aspect}' does not exist.")
        if la_str not in a.levels:
            raise ValueError(f"Aspect level '{la}' [{a_type}] does not exist.")
        if lb_str not in a.levels:
            raise ValueError(f"Aspect level '{lb}' [{a_type}] does not exist.")

        # Step 1: compute current closure as the staging area
        staged, _, _ = self.closure()

        # Step 2: apply the requested addition to the staging area
        zero  = NATURAL_ZERO
        vd_ab = VDiff(aspect, la_str, lb_str)
        vd_ba = VDiff(aspect, lb_str, la_str)
        origin = ['SETREL', [aspect, la_str, rel, lb_str]]
        staged_adds, staged_colls = [], []
        if rel == UNDEFINED:
            app_ac(origin, set_vdiff_relation(staged, vd_ab, zero, UNDEFINED), staged_adds, staged_colls)
            app_ac(origin, set_vdiff_relation(staged, vd_ba, zero, UNDEFINED), staged_adds, staged_colls)
            app_ac(origin, set_vdiff_relation(staged, zero, vd_ab, UNDEFINED), staged_adds, staged_colls)
            app_ac(origin, set_vdiff_relation(staged, zero, vd_ba, UNDEFINED), staged_adds, staged_colls)
        elif rel == BT:
            app_ac(origin, set_vdiff_relation(staged, vd_ab, zero, TRUE),  staged_adds, staged_colls)
            app_ac(origin, set_vdiff_relation(staged, vd_ba, zero, FALSE), staged_adds, staged_colls)
            app_ac(origin, set_vdiff_relation(staged, zero, vd_ab, FALSE), staged_adds, staged_colls)
            app_ac(origin, set_vdiff_relation(staged, zero, vd_ba, TRUE),  staged_adds, staged_colls)
        elif rel == BTE:
            app_ac(origin, set_vdiff_relation(staged, vd_ab, zero, TRUE),  staged_adds, staged_colls)
            app_ac(origin, set_vdiff_relation(staged, zero, vd_ba, TRUE),  staged_adds, staged_colls)
        elif rel == EQ:
            app_ac(origin, set_vdiff_relation(staged, vd_ab, zero, TRUE),  staged_adds, staged_colls)
            app_ac(origin, set_vdiff_relation(staged, vd_ba, zero, TRUE),  staged_adds, staged_colls)
            app_ac(origin, set_vdiff_relation(staged, zero, vd_ab, TRUE),  staged_adds, staged_colls)
            app_ac(origin, set_vdiff_relation(staged, zero, vd_ba, TRUE),  staged_adds, staged_colls)
        elif rel == WTE:
            app_ac(origin, set_vdiff_relation(staged, vd_ba, zero, TRUE),  staged_adds, staged_colls)
            app_ac(origin, set_vdiff_relation(staged, zero, vd_ab, TRUE),  staged_adds, staged_colls)
        elif rel == WT:
            app_ac(origin, set_vdiff_relation(staged, vd_ba, zero, TRUE),  staged_adds, staged_colls)
            app_ac(origin, set_vdiff_relation(staged, vd_ab, zero, FALSE), staged_adds, staged_colls)
            app_ac(origin, set_vdiff_relation(staged, zero, vd_ba, FALSE), staged_adds, staged_colls)
            app_ac(origin, set_vdiff_relation(staged, zero, vd_ab, TRUE),  staged_adds, staged_colls)

        # Step 3: immediate collision in the staged area — reject
        if staged_colls:
            return ([], staged_colls, [])

        # Step 4-5: recompute closure on the staged dict and check for inferred collisions
        # Temporarily swap the matrix for the staged dict to reuse self.closure()
        original_matrix = self.vdiff_comparison_matrix
        self.vdiff_comparison_matrix = staged
        _, inferred_adds, inferred_colls = self.closure()
        self.vdiff_comparison_matrix = original_matrix

        # Step 5: inferred collision — reject
        if inferred_colls:
            return ([], inferred_colls, [])

        # Step 6: clean — commit only the explicit addition to the real matrix
        adds, colls = [], []
        if rel == UNDEFINED:
            app_ac(origin, set_vdiff_relation(original_matrix, vd_ab, zero, UNDEFINED), adds, colls)
            app_ac(origin, set_vdiff_relation(original_matrix, vd_ba, zero, UNDEFINED), adds, colls)
            app_ac(origin, set_vdiff_relation(original_matrix, zero, vd_ab, UNDEFINED), adds, colls)
            app_ac(origin, set_vdiff_relation(original_matrix, zero, vd_ba, UNDEFINED), adds, colls)
        elif rel == BT:
            app_ac(origin, set_vdiff_relation(original_matrix, vd_ab, zero, TRUE),  adds, colls)
            app_ac(origin, set_vdiff_relation(original_matrix, vd_ba, zero, FALSE), adds, colls)
            app_ac(origin, set_vdiff_relation(original_matrix, zero, vd_ab, FALSE), adds, colls)
            app_ac(origin, set_vdiff_relation(original_matrix, zero, vd_ba, TRUE),  adds, colls)
        elif rel == BTE:
            app_ac(origin, set_vdiff_relation(original_matrix, vd_ab, zero, TRUE),  adds, colls)
            app_ac(origin, set_vdiff_relation(original_matrix, zero, vd_ba, TRUE),  adds, colls)
        elif rel == EQ:
            app_ac(origin, set_vdiff_relation(original_matrix, vd_ab, zero, TRUE),  adds, colls)
            app_ac(origin, set_vdiff_relation(original_matrix, vd_ba, zero, TRUE),  adds, colls)
            app_ac(origin, set_vdiff_relation(original_matrix, zero, vd_ab, TRUE),  adds, colls)
            app_ac(origin, set_vdiff_relation(original_matrix, zero, vd_ba, TRUE),  adds, colls)
        elif rel == WTE:
            app_ac(origin, set_vdiff_relation(original_matrix, vd_ba, zero, TRUE),  adds, colls)
            app_ac(origin, set_vdiff_relation(original_matrix, zero, vd_ab, TRUE),  adds, colls)
        elif rel == WT:
            app_ac(origin, set_vdiff_relation(original_matrix, vd_ba, zero, TRUE),  adds, colls)
            app_ac(origin, set_vdiff_relation(original_matrix, vd_ab, zero, FALSE), adds, colls)
            app_ac(origin, set_vdiff_relation(original_matrix, zero, vd_ba, FALSE), adds, colls)
            app_ac(origin, set_vdiff_relation(original_matrix, zero, vd_ab, TRUE),  adds, colls)
        return (adds, [], inferred_adds)

    def get_aspect_level_relation(self, aspect: str, la, lb) -> str:
        a = self.get_aspect(aspect)
        a_type = a.data_type
        la_str, lb_str = str(la), str(lb)
        if a is None:
            raise ValueError(f"Aspect '{aspect}' does not exist.")
        if not la_str in a.levels:
            raise ValueError(f"Aspect level '{la}' [{a_type}] does not exist.")
        if not lb_str in a.levels:
            raise ValueError(f"Aspect level '{lb}' [{a_type}] does not exist.")
        zero = NATURAL_ZERO
        vd_ab = VDiff(aspect, la_str, lb_str)
        rel_ab_z = self.get_vdiff_relation(vd_ab, zero)
        rel_z_ab = self.get_vdiff_relation(zero, vd_ab)
        if rel_ab_z == TRUE and rel_z_ab == FALSE:
            return BT
        elif rel_ab_z == TRUE and rel_z_ab == UNDEFINED:
            return BTE
        elif rel_ab_z == TRUE and rel_z_ab == TRUE:
            return EQ
        elif rel_ab_z == UNDEFINED and rel_z_ab == TRUE:
            return WTE
        elif rel_ab_z == FALSE and rel_z_ab == TRUE:
            return WT
        elif rel_ab_z == UNDEFINED and rel_z_ab == UNDEFINED:
            return UNDEFINED
        else:
            return NotImplemented

    def get_vdiff_relation(self, vd1: VDiff, vd2: VDiff) -> str:
        return get_vdiff_relation(self.vdiff_comparison_matrix, vd1, vd2)
    
    def try_set_vdiff_order_relation(self,
                                      vd1: VDiff, vd2: VDiff,
                                      order_rel: str) -> Tuple:
        """Validate and commit a vdiff order relation using a full VDCM staging copy.

        order_rel must be one of: GT ('⊐'), GTE ('⊒'), DEQ ('≜'),
                                   LTE ('⊑'), LT ('⊏'), UNDEFINED ('').

        Mapping to VDCM TRUE/FALSE entries (via set_rel):
          ⊐  →  vd1⊒vd2  AND  vd2⋣vd1
          ⊒  →  vd1⊒vd2
          ≜  →  vd1⊒vd2  AND  vd2⊒vd1  (and mirror pairs)
          ⊑  →  vd2⊒vd1
          ⊏  →  vd2⊒vd1  AND  vd1⋣vd2
          —  →  unset all four VDCM entries for (vd1,vd2)

        Flow (mirrors try_set_aspect_level_relation):
          1. Deep-copy the full VDCM as staging area.
          2. Apply set_rel to staging. Reject on immediate collision.
          3. Recompute closure on staging. Reject on inferred collision.
          4. If clean, commit set_rel to the real VDCM.
             Return (adds, [], inferred_adds).
        """
        import copy

        # Decompose vd1/vd2 into (aspect, la, lb) for set_rel
        an1, l1a, l1b = vd1.aspect_name, vd1.from_level, vd1.to_level
        an2, l2a, l2b = vd2.aspect_name, vd2.from_level, vd2.to_level

        origin = ['SETVDREL', [repr(vd1), order_rel, repr(vd2)]]

        # ── Step 1: deep-copy the VDCM ──────────────────────────
        staged_matrix = copy.deepcopy(self.vdiff_comparison_matrix)

        # ── Step 2: apply to staging ─────────────────────────────
        staged_adds, staged_colls = [], []
        original_matrix = self.vdiff_comparison_matrix
        self.vdiff_comparison_matrix = staged_matrix

        if order_rel == UNDEFINED:
            # Unset all four entries for this vdiff pair
            for va, vb in [(vd1, vd2), (vd2, vd1)]:
                app_ac(origin,
                       set_vdiff_relation(staged_matrix, va, vb, UNDEFINED),
                       staged_adds, staged_colls)
                app_ac(origin,
                       set_vdiff_relation(staged_matrix, vb, va, UNDEFINED),
                       staged_adds, staged_colls)
        else:
            s_adds, s_colls = self.set_rel(an1, l1a, l1b, an2, l2a, l2b, order_rel)
            staged_adds.extend(s_adds)
            staged_colls.extend(s_colls)

        self.vdiff_comparison_matrix = original_matrix

        # ── Step 3: immediate collision — reject ─────────────────
        if staged_colls:
            return ([], staged_colls, [])

        # ── Step 4: recompute closure on staging, check inferred ─
        self.vdiff_comparison_matrix = staged_matrix
        _, inferred_adds, inferred_colls = self.closure()
        self.vdiff_comparison_matrix = original_matrix

        if inferred_colls:
            return ([], inferred_colls, [])

        # ── Step 5: clean — commit to the real VDCM ─────────────
        adds, colls = [], []
        if order_rel == UNDEFINED:
            for va, vb in [(vd1, vd2), (vd2, vd1)]:
                app_ac(origin,
                       set_vdiff_relation(original_matrix, va, vb, UNDEFINED),
                       adds, colls)
                app_ac(origin,
                       set_vdiff_relation(original_matrix, vb, va, UNDEFINED),
                       adds, colls)
        else:
            c_adds, c_colls = self.set_rel(an1, l1a, l1b, an2, l2a, l2b, order_rel)
            adds.extend(c_adds)
            colls.extend(c_colls)

        return (adds, [], inferred_adds)

    def set_vdiff_relation(self, vd1: VDiff, vd2: VDiff, new_rel: str) -> Tuple:
        return set_vdiff_relation(self.vdiff_comparison_matrix, vd1, vd2, new_rel)

    def set_rel(self, a1: str, l1a, l1b, a2, l2a, l2b, rel: str) -> Tuple:
        adds, colls = [], []
        a1_ab = VDiff(a1, l1a, l1b)
        a1_ba = VDiff(a1, l1b, l1a)
        a2_ab = VDiff(a2, l2a, l2b)
        a2_ba = VDiff(a2, l2b, l2a)
        origin = ['SETREL', [a1_ab, rel, a2_ab]]
        if rel == GT:
            app_ac(origin, self.set_vdiff_relation(a1_ab, a2_ab, TRUE), adds, colls)
            app_ac(origin, self.set_vdiff_relation(a2_ab, a1_ab, FALSE), adds, colls)
        elif rel == GTE:
            app_ac(origin, self.set_vdiff_relation(a1_ab, a2_ab, TRUE), adds, colls)
        elif rel == DEQ:
            app_ac(origin, self.set_vdiff_relation(a1_ab, a2_ab, TRUE), adds, colls)
            app_ac(origin, self.set_vdiff_relation(a2_ab, a1_ab, TRUE), adds, colls)
            app_ac(origin, self.set_vdiff_relation(a1_ba, a2_ba, TRUE), adds, colls)
            app_ac(origin, self.set_vdiff_relation(a2_ba, a1_ba, TRUE), adds, colls)
        elif rel == LTE:
            app_ac(origin, self.set_vdiff_relation(a2_ab, a1_ab, TRUE), adds, colls)
        elif rel == LT:
            app_ac(origin, self.set_vdiff_relation(a2_ab, a1_ab, TRUE), adds, colls)
            app_ac(origin, self.set_vdiff_relation(a1_ab, a2_ab, FALSE), adds, colls)
        return (adds, colls)

    def pos(self, an: str, la, lb) -> bool:
        # TODO: Error handling
        return pos(VDiff(an, la, lb), self.aspects[an], self.vdiff_comparison_matrix)

    def non_neg(self, an: str, la, lb) -> bool:
        # TODO: Error handling
        return non_neg(VDiff(an, la, lb), self.aspects[an], self.vdiff_comparison_matrix)

    def zero(self, an: str, la, lb) -> bool:
        # TODO: Error handling
        return zero(VDiff(an, la, lb), self.aspects[an], self.vdiff_comparison_matrix)

    def non_pos(self, an: str, la, lb) -> bool:
        # TODO: Error handling
        return non_pos(VDiff(an, la, lb), self.aspects[an], self.vdiff_comparison_matrix)

    def neg(self, an: str, la, lb) -> bool:
        # TODO: Error handling
        return neg(VDiff(an, la, lb), self.aspects[an], self.vdiff_comparison_matrix)
    
    def vd_enum_verbose(self):
        for a in self.aspects.values():
            for l_from in a.levels:
                for l_to in a.levels:
                    yield VDiff(a.name, l_from, l_to)

    def vd_enum_brief(self):
        for a in self.aspects.values():
            for vd in a.vdiffs:
                yield vd

    def vdc_enum(self):
        """Iterate every (vd1, vd2, rel) triple stored in the vdcm."""
        for vd1, row in self.vdiff_comparison_matrix.items():
            for vd2, rel in row.items():
                yield (vd1, vd2, rel)

    def closure(self):
        adds, colls = [], []
        # Shallow copy: each inner row dict is a new dict so writes don't touch self.vdiff_comparison_matrix
        closure = {vd1: dict(row) for vd1, row in self.vdiff_comparison_matrix.items()}
        # Compute closure
        prev_adds = -1
        while len(adds) != prev_adds:
            prev_adds = len(adds)
            # Phase 1: DiffP / NegDiffP (same-aspect only)
            for asp_name, asp in self.aspects.items():
                for c in asp.levels:
                    for d in asp.levels:
                        cd = VDiff(asp_name, c, d)
                        for e in asp.levels:
                            for f in asp.levels:
                                ef = VDiff(asp_name, e, f)
                                rel_cd_ef = get_vdiff_relation(closure, cd, ef)
                                if rel_cd_ef == TRUE: # cd⊒ef => ce⊒df
                                    ce = VDiff(asp_name, c, e)
                                    df = VDiff(asp_name, d, f)
                                    origin = ['DiffP', [cd, rel_cd_ef, ef]]
                                    app_ac(origin, set_vdiff_relation(closure, ce, df, TRUE), adds, colls)
                                elif rel_cd_ef == FALSE: # cd⋣ef => fd⋣ec
                                    fd = VDiff(asp_name, f, d)
                                    ec = VDiff(asp_name, e, c)
                                    origin = ['NegDiffP', [fd, rel_cd_ef, ec]]
                                    app_ac(origin, set_vdiff_relation(closure, fd, ec, FALSE), adds, colls)
                                if colls: # A collision has occurred — abort
                                    return (closure, adds, colls)
            # Phase 2: TransP / InvP / NegTransP / NegInvP (cd as pivot)
            for cd in self.vd_enum_verbose():
                for ab in self.vd_enum_verbose():
                    rel_ab_cd = get_vdiff_relation(closure, ab, cd)
                    if rel_ab_cd == TRUE: # ab⊒cd
                        for ef in self.vd_enum_verbose():
                            rel_cd_ef = get_vdiff_relation(closure, cd, ef)
                            if rel_cd_ef == TRUE: # ab⊒cd & cd⊒ef ==> ab⊒ef
                                origin = ['TransP', [ab, rel_ab_cd, cd, rel_cd_ef, ef]]
                                app_ac(origin, set_vdiff_relation(closure, ab, ef, TRUE), adds, colls)
                                if ef.natural_zero(): # ab⊒cd & cd⊒xx ==> dc⊒ba
                                    ba = ab.inv()
                                    dc = cd.inv()
                                    origin = ['InvP_R', [ab, rel_ab_cd, cd, rel_cd_ef, ef]]
                                    app_ac(origin, set_vdiff_relation(closure, dc, ba, TRUE), adds, colls)
                                if ab.natural_zero(): # xx⊒cd & cd⊒ef ==> fe⊒dc
                                    fe = ef.inv()
                                    dc = cd.inv()
                                    origin = ['InvP_L', [ab, rel_ab_cd, cd, rel_cd_ef, ef]]
                                    app_ac(origin, set_vdiff_relation(closure, fe, dc, TRUE), adds, colls)
                            elif rel_cd_ef == FALSE: # ab⊒cd & cd⋣ef
                                rel_cd_ab = get_vdiff_relation(closure, cd, ab)
                                if rel_cd_ab == TRUE: # ab≜cd & cd⋣ef ==> ab⋣ef
                                    origin = ['NegTransP_DEQ_L', [ab, DEQ, cd, FALSE, ef]]
                                    app_ac(origin, set_vdiff_relation(closure, ab, ef, FALSE), adds, colls)
                            if colls: # A collision has occurred — abort
                                return (closure, adds, colls)
                    elif rel_ab_cd == FALSE: # ab⋣cd
                        for ef in self.vd_enum_verbose():
                            rel_cd_ef = get_vdiff_relation(closure, cd, ef)
                            if rel_cd_ef == TRUE: # ab⋣cd & cd⊒ef
                                rel_ef_cd = get_vdiff_relation(closure, ef, cd)
                                if rel_ef_cd == TRUE: # ab⋣cd & cd≜ef ==> ab⋣ef
                                    origin = ['NegTransP_DEQ_R', [ab, FALSE, cd, DEQ, ef]]
                                    app_ac(origin, set_vdiff_relation(closure, ab, ef, FALSE), adds, colls)
                            elif rel_cd_ef == FALSE: # ab⋣cd & cd⋣ef ==> ab⋣ef
                                origin = ['NegTransP', [ab, rel_ab_cd, cd, rel_cd_ef, ef]]
                                app_ac(origin, set_vdiff_relation(closure, ab, ef, FALSE), adds, colls)
                                if ab.natural_zero(): # xx⋣cd & cd⋣ef ==> fe⋣dc
                                    dc = cd.inv()
                                    fe = ef.inv()
                                    origin = ['NegInvP_L', [ab, rel_ab_cd, cd, rel_cd_ef, ef]]
                                    app_ac(origin, set_vdiff_relation(closure, fe, dc, FALSE), adds, colls)
                                if ef.natural_zero(): # ab⋣cd & cd⋣xx ==> dc⋣ba
                                    ba = ab.inv()
                                    dc = cd.inv()
                                    origin = ['NegInvP_R', [ab, rel_ab_cd, cd, rel_cd_ef, ef]]
                                    app_ac(origin, set_vdiff_relation(closure, dc, ba, FALSE), adds, colls)
                            if colls: # A collision has occurred — abort
                                return (closure, adds, colls)
        return (closure, adds, colls)

    def expand_vdiff_comparison_matrix(self, an2: str):
        """Add vdcm entries for all new vdiff pairs that involve aspect *an2*.

        Called whenever a level is added to an aspect.  Iterates every existing
        aspect (including an2 itself) and cross-products its vdiffs with those
        of an2.  Entries are only written when absent, so existing relations are
        never overwritten.

        Diagonal initialisation: k1 == k2 → TRUE (covers both the same non-zero
        vdiff and the NATURAL_ZERO × NATURAL_ZERO self-pair).
        """
        vdcm = self.vdiff_comparison_matrix
        for a1 in self.aspects.values():
            for vd1 in a1.vdiffs:
                k1 = _vdiff_key(vd1)
                for vd2 in self.aspects[an2].vdiffs:
                    k2 = _vdiff_key(vd2)
                    rel = TRUE if k1 == k2 else UNDEFINED
                    # Forward: k1 → k2
                    if k1 not in vdcm:
                        vdcm[k1] = {}
                    if k2 not in vdcm[k1]:
                        logger.debug(f"Initialising {k1}?{k2}: {rel}")
                        vdcm[k1][k2] = rel
                    # Reverse: k2 → k1
                    if k2 not in vdcm:
                        vdcm[k2] = {}
                    if k1 not in vdcm[k2]:
                        logger.debug(f"Initialising {k2}?{k1}: {rel}")
                        vdcm[k2][k1] = rel
    
    def compute_consequence_space(self) -> List:
        """Derive the full consequence space from aspects and their levels.
        This is always computable from first principles and need not be persisted.
        Aspects with no levels contribute a single None placeholder so the
        table remains non-empty while the aspect is being populated."""
        aspects = list(self.aspects.values())
        if not aspects:
            return [Consequence()]
        level_lists = [list(a.levels.keys()) or [None] for a in aspects]
        result = []
        for combo in product(*level_lists):
            c = Consequence({a.name: level for a, level in zip(aspects, combo)})
            result.append(c)
        return result

    @property
    def consequence_space(self) -> List:
        """Computed on demand from aspects and levels — never persisted."""
        return self.compute_consequence_space()

    def export_aspect_to_excel(self, aspect_name: str, filename: str) -> int:
        import openpyxl
        title = ASP + aspect_name
        try:
            wb = openpyxl.load_workbook(filename)
        except: # Create new workbook if file does not exist
            wb = openpyxl.Workbook()
            ws_a = wb.active
            ws_a.title = title
        try:
            ws_a = wb[title]
        except: # Create new sheet if it does not exist
            ws_a = wb.create_sheet(title=title)
        aspect = self.aspects[aspect_name]
        n_levels = self.export_aspect_to_worksheet(aspect, ws_a)
        wb.save(filename)
        return n_levels

    def reorder_aspects(self, new_order: list) -> None:
        """Reorder self.aspects to match new_order.
        new_order must contain exactly the same names as self.aspects.
        Raises ValueError if they don't match.
        """
        existing = set(self.aspects.keys())
        requested = list(new_order)
        if set(requested) != existing:
            missing  = existing - set(requested)
            extra    = set(requested) - existing
            msgs = []
            if missing:  msgs.append(f"missing from new order: {sorted(missing)}")
            if extra:    msgs.append(f"unknown aspects: {sorted(extra)}")
            raise ValueError("; ".join(msgs))
        self.aspects = {name: self.aspects[name] for name in requested}

    def export_proj_tab_to_worksheet(self, ws,
                                      project_name: str = "",
                                      author: str = "") -> None:
        """Write the |PROJ| tab to ws."""
        ws["A1"] = "EUDOXA"
        ws["B1"] = "0.1"
        ws["A2"] = "Project name:"
        ws["B2"] = project_name
        ws["A3"] = "Author:"
        ws["B3"] = author
        ws["A4"] = "Aspects:"
        for i, asp_name in enumerate(self.aspects.keys(), start=5):
            ws.cell(row=i, column=1).value = "-"
            ws.cell(row=i, column=2).value = asp_name

    def export_project_to_workbook(self):
        """Export the full project to a new openpyxl Workbook.
        Writes one |ASP| tab per aspect (levels + relations),
        one |CONS| tab, and one |VDCM| tab.
        Returns the workbook object.
        """
        import openpyxl
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove the default empty sheet

        # |PROJ| tab first
        ws_proj = wb.create_sheet(title=PROJ)
        self.export_proj_tab_to_worksheet(
            ws_proj,
            project_name=getattr(self, "project_name", ""),
            author=getattr(self, "author", "")
        )

        # One |ASP| tab per aspect
        for asp_name, aspect in self.aspects.items():
            ws = wb.create_sheet(title=ASP + asp_name)
            self.export_aspect_to_worksheet(aspect, ws)
            self.export_aspect_level_relations_to_worksheet(aspect, ws)

        # |CONS| tab
        ws_cons = wb.create_sheet(title=CONS)
        self.export_consequences_to_worksheet(ws_cons)

        # |VDCM| tab
        ws_vdcm = wb.create_sheet(title=VDCM)
        self.export_vdiff_comparison_matrix_to_worksheet(
            self.vdiff_comparison_matrix, ws_vdcm
        )

        return wb

    def export_aspect_level_relations_to_excel(self, aspect_name: str, filename: str):
        import openpyxl
        title = ASP + aspect_name
        try:
            wb = openpyxl.load_workbook(filename)
        except: # Create new workbook if file does not exist
            wb = openpyxl.Workbook()
            ws_a = wb.active
            ws_a.title = title
        try:
            ws_a = wb[title]
        except: # Create new sheet if it does not exist
            ws_a = wb.create_sheet(title=title)
        aspect = self.aspects[aspect_name]
        self.export_aspect_level_relations_to_worksheet(aspect, ws_a)
        wb.save(filename)

    def export_consequences_to_excel(self, filename: str):
        import openpyxl
        try:
            wb = openpyxl.load_workbook(filename)
        except: # Create new workbook if file does not exist
            wb = openpyxl.Workbook()
            ws_c = wb.active
            ws_c.title = CONS
        try:
            ws_c = wb[CONS]
        except: # Create new sheet if it does not exist
            ws_c = wb.create_sheet(title=CONS)
        self.export_consequences_to_worksheet(ws_c)
        wb.save(filename)

    def export_aspect_to_worksheet(self, aspect, ws) -> int:
        start_row = 3
        # Row 1: aspect name and data type
        ws["A1"] = aspect.name
        ws["B1"] = aspect.data_type.__name__

        # Row 2: aspect description
        ws["A2"] = aspect.description

        # From row 3: levels (name and description)
        # TODO: Handle data types
        row_index = start_row
        for level_str, description in aspect.levels.items():
            level = parse_type(level_str, aspect.data_type)
            ws.cell(row=row_index, column=1).value=level
            ws.cell(row=row_index, column=2).value=description
            row_index += 1
        return row_index-start_row

    def export_aspect_level_relations_to_worksheet(self, aspect, ws):
        # Clear all non-empty cells from column 5 onwards
        max_row = ws.max_row
        max_col = ws.max_column
        for col in range(5, max_col + 1):
            for row in range(1, max_row + 1):
                ws.cell(row=row, column=col).value=None
        # Row 2: column headers
        col_index = 5
        for level_str in aspect.levels:
            level = parse_type(level_str, aspect.data_type)
            ws.cell(row=2, column=col_index).value=level
            col_index += 1
      
        # From row 3: row headers and values
        row_index = 3
        for ls_row in aspect.levels:
            l_row = parse_type(ls_row, aspect.data_type)
            ws.cell(row=row_index, column=4).value=l_row
            col_index = 5
            for ls_col in aspect.levels:
                rel = self.get_aspect_level_relation(aspect.name, ls_row, ls_col)
                rel = None if rel == UNDEFINED else rel
                logger.debug(f"Exporting {ls_row} {rel} {ls_col} in '{aspect.name}'")
                ws.cell(row=row_index, column=col_index).value=rel
                col_index +=1
            row_index += 1

    def _vd_label(self, vd):
        """Return the label for a VDiff used in Excel and display."""
        if vd.natural_zero():
            return ZDIFF_DISPLAY
        return f"({vd.from_level},{vd.to_level})"

    def _sorted_vdiffs(self, asp):
        """Return asp.vdiffs sorted: zero-diff first, then (from, to) pairs
        sorted by (level_index[from], level_index[to])."""
        level_order = {l: i for i, l in enumerate(asp.levels.keys())}
        zero = [vd for vd in asp.vdiffs if vd.natural_zero()]
        rest = [vd for vd in asp.vdiffs if not vd.natural_zero()]
        rest.sort(key=lambda vd: (level_order.get(vd.from_level, 0),
                                  level_order.get(vd.to_level, 0)))
        return zero + rest

    def export_vdiff_comparison_matrix_to_worksheet(self, vdcm, ws):
        """Write the full vdiff comparison matrix to a worksheet."""

        # Corner cell
        ws.cell(row=3, column=3).value = f"{DELTA}\\{DELTA}"

        # Row 2+3: column headers
        col_index = 4
        for an, asp in self.aspects.items():
            ws.cell(row=2, column=col_index).value = an
            for vd in self._sorted_vdiffs(asp):
                ws.cell(row=3, column=col_index).value = self._vd_label(vd)
                col_index += 1

        # Row 4+: row headers and cell values
        ordered = [(an, vd) for an, asp in self.aspects.items()
                   for vd in self._sorted_vdiffs(asp)]
        prev_an1 = None
        for row_index, (an1, vd1) in enumerate(ordered, start=4):
            if an1 != prev_an1:
                ws.cell(row=row_index, column=2).value = an1
                prev_an1 = an1
            ws.cell(row=row_index, column=3).value = self._vd_label(vd1)
            for col_offset, (an2, vd2) in enumerate(ordered):
                rel = get_vdiff_relation(self.vdiff_comparison_matrix, vd1, vd2)
                ws.cell(row=row_index, column=4 + col_offset).value = rel

    def export_vdiff_comparison_matrix_to_excel(self, filename: str):
        import openpyxl
        title = VDCM
        try:
            wb = openpyxl.load_workbook(filename)
        except: # Create new workbook if file does not exist
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = title
        try:
            ws = wb[title]
        except: # Create new sheet if it does not exist
            ws = wb.create_sheet(title=title)
        self.export_vdiff_comparison_matrix_to_worksheet(self.vdiff_comparison_matrix, ws)
        wb.save(filename)
                
    def import_proj_tab_from_worksheet(self, ws) -> dict:
        """Parse the |PROJ| tab. Returns:
        {
          'project_name': str,
          'author': str,
          'aspect_order': [str, ...],   # may be empty if section absent
        }
        """
        result = {"project_name": "", "author": "", "aspect_order": []}
        in_aspects = False
        for row in ws.iter_rows(min_row=2, values_only=True):
            key = str(row[0]).strip() if row[0] is not None else ""
            val = str(row[1]).strip() if row[1] is not None else ""
            if key == "Project name:":
                result["project_name"] = val
                in_aspects = False
            elif key == "Author:":
                result["author"] = val
                in_aspects = False
            elif key == "Aspects:":
                in_aspects = True
            elif in_aspects and key == "-" and val:
                result["aspect_order"].append(val)
            elif key and key != "-":
                in_aspects = False
        return result

    def import_vdiff_comparison_matrix_from_worksheet(self, ws) -> dict:
        """Import vdiff relations from a |VDCM| worksheet.

        Layout matches export_vdiff_comparison_matrix_to_worksheet:
          Row 2: aspect name column headers
          Row 3: vdiff label column headers ('◬' or '(0,600)')
          Col B: aspect name row headers
          Col C: vdiff label row headers
          Cells D4+: relation values (TRUE='⋒', FALSE='⋓', or empty)

        Returns {"adds": [...], "collisions": [...]}.
        """
        def parse_label(lbl):
            lbl = lbl.strip()
            if lbl == ZDIFF_DISPLAY:
                return ZDIFF_TUPLE
            lbl = lbl.strip("()")
            parts = lbl.split(",", 1)
            if len(parts) != 2:
                return None
            a, b = parts[0].strip(), parts[1].strip()
            return (a, b)

        # Parse column headers (row 3, from col D)
        col_headers = []   # [(aspect_name, vd_key_tuple), ...]
        current_asp = None
        col = 4
        while True:
            asp_val = ws.cell(row=2, column=col).value
            vd_val  = ws.cell(row=3, column=col).value
            if asp_val is None and vd_val is None:
                break
            if asp_val is not None:
                current_asp = str(asp_val).strip()
            if vd_val is not None and current_asp is not None:
                d = parse_label(str(vd_val))
                if d is not None:
                    col_headers.append((current_asp, d))
            col += 1

        adds       = []
        collisions = []

        # Parse data rows (row 4 onwards)
        # col B has aspect name only on first row of each block — track it
        current_row_asp = None
        row = 4
        while True:
            row_asp = ws.cell(row=row, column=2).value
            row_lbl = ws.cell(row=row, column=3).value
            if row_asp is None and row_lbl is None:
                break
            if row_asp is not None:
                current_row_asp = str(row_asp).strip()
            if row_lbl is None or current_row_asp is None:
                row += 1
                continue
            an1 = current_row_asp
            d1  = parse_label(str(row_lbl).strip())
            if d1 is None or an1 not in self.aspects:
                row += 1
                continue
            vd1 = VDiff(an1, None, None) if d1 == ZDIFF_TUPLE \
                  else VDiff(an1, d1[0], d1[1])
            for col_offset, (an2, d2) in enumerate(col_headers):
                if an2 not in self.aspects:
                    continue
                cell_val = ws.cell(row=row, column=4 + col_offset).value
                if cell_val is None or str(cell_val).strip() == "":
                    continue
                new_rel = str(cell_val).strip()
                if new_rel not in (TRUE, FALSE):
                    continue
                vd2 = VDiff(an2, None, None) if d2 == ZDIFF_TUPLE \
                      else VDiff(an2, d2[0], d2[1])
                add, coll = set_vdiff_relation(
                    self.vdiff_comparison_matrix, vd1, vd2, new_rel
                )
                if add:  adds.append(add)
                if coll: collisions.append(coll)
            row += 1

        return {"adds": adds, "collisions": collisions}

    def export_consequences_to_worksheet(self, ws):
        # Row 1: headers, Row 2: data types
        for col_index, aspect_name in enumerate(self.aspects.keys(), start=2):
            ws.cell(row=1, column=col_index).value=aspect_name
            aspect_type = self.aspects[aspect_name].data_type
            ws.cell(row=2, column=col_index).value=aspect_type.__name__

        # Consequences (from row 3)
        for row_index, (short_name, consequence) in enumerate(self.consequences.items(), start=3):
            ws.cell(row=row_index, column=1).value=short_name
            for col_index, aspect_name in enumerate(self.aspects.keys(), start=2):
                aspect_type = self.aspects[aspect_name].data_type
                level_str = consequence[aspect_name]
                level = None if level_str is None else parse_type(level_str, aspect_type)
                ws.cell(row=row_index, column=col_index).value=level

    def import_aspect_from_excel(self, aspect_name: str, filename: str):
        import openpyxl
        wb = openpyxl.load_workbook(filename, data_only=True)
        ws = wb[ASP + aspect_name]
        return self.import_aspect_from_worksheet(ws)

    def import_aspect_level_relations_from_excel(self, aspect_name: str, filename: str):
        import openpyxl
        wb = openpyxl.load_workbook(filename, data_only=True)
        ws = wb[ASP + aspect_name]
        return self.import_aspect_level_relations_from_worksheet(ws)

    def import_consequences_from_excel(self, filename: str):
        import openpyxl
        wb = openpyxl.load_workbook(filename, data_only=True)
        ws = wb[CONS]
        self.import_consequences_from_worksheet(ws)

    def validate_and_import_workbook(self, wb, base_mgr=None) -> dict:
        """Staged validate-then-commit import from an openpyxl workbook.

        Pipeline:
          1. Per-aspect: add aspect, levels, and relations to a temporary
             manager. Collect all per-aspect errors; cancel if any.
          2. Compute closure on the temporary manager. Cancel on collision.
          3. Import named consequences (3a strict). Cancel on any error.
          4. On full success, commit by copying temporary state into self.

        base_mgr: optional EudoxaManager to use as the starting state of the
                  temporary manager (for single-aspect import into an existing
                  project). Defaults to a fresh empty manager.
        """
        result = {
            "success":              False,
            "project_name":         "",
            "author":               "",
            "imported_aspects":     [],
            "aspect_errors":        {},
            "skipped_asp_tabs":     [],
            "missing_asp_tabs":     [],
            "closure_collisions":   [],
            "vdcm_adds":            0,
            "vdcm_add_details":     [],
            "imported_consequences": [],
            "imported_consequence_details": [],
            "consequence_errors":   [],
            "missing_cons_sheet":   False,
        }

        # Initialise the temporary manager
        if base_mgr is not None:
            tmp = EudoxaManager.from_dict(base_mgr.to_dict())
        else:
            tmp = EudoxaManager()

        # ── Step 0: parse |PROJ| tab if present ────────────────
        proj_sheet = next((n for n in wb.sheetnames if n == PROJ), None)
        aspect_order = []   # canonical order from |PROJ|; empty = use tab order
        if proj_sheet is not None:
            proj_info = tmp.import_proj_tab_from_worksheet(wb[proj_sheet])
            result["project_name"] = proj_info["project_name"]
            result["author"]       = proj_info["author"]
            aspect_order           = proj_info["aspect_order"]
            # Check for listed aspects with missing |ASP| tabs
            available_tabs = {n[len(ASP):] for n in wb.sheetnames
                              if n.startswith(ASP)}
            missing_tabs = [a for a in aspect_order if a not in available_tabs]
            if missing_tabs:
                result["missing_asp_tabs"] = missing_tabs
                return result
            # Note any |ASP| tabs not listed in |PROJ| (will be skipped)
            result["skipped_asp_tabs"] = [
                a for a in available_tabs if a not in aspect_order
            ]

        # ── Step 1: aspects, levels, and relations ─────────────────
        prefix = ASP  # '|ASP| '
        if aspect_order:
            # Use |PROJ| order; skip tabs not in the list
            aspect_sheets = [ASP + name for name in aspect_order
                             if ASP + name in wb.sheetnames]
        else:
            aspect_sheets = [name for name in wb.sheetnames
                             if name.startswith(prefix)]

        for sheet_name in aspect_sheets:
            ws = wb[sheet_name]
            aspect_name = ws["A1"].value
            errors = []

            # Add aspect and levels (stops on first error per aspect)
            try:
                tmp.import_aspect_from_worksheet(ws)
            except ValueError as e:
                errors.append(str(e))
                result["aspect_errors"][aspect_name] = errors
                continue

            # Add relations if present
            has_relations = ws.cell(row=2, column=5).value is not None
            if has_relations:
                adds, colls = tmp.import_aspect_level_relations_from_worksheet(ws)
                if colls:
                    for coll_entry in colls:
                        _, _, coll = coll_entry
                        vd1, old_rel, vd2, new_rel = coll
                        errors.append(
                            f"Collision: attempted {repr(vd1)} {new_rel} {repr(vd2)} "
                            f"conflicts with existing {repr(vd1)} {old_rel} {repr(vd2)}"
                        )
                    result["aspect_errors"][aspect_name] = errors
                    continue

            asp_obj = tmp.aspects[aspect_name]
            # Collect non-trivial relations for detail view
            # Only show BT/BTE/EQ to avoid showing both A≻B and B≺A
            rel_details = []
            for la in asp_obj.levels:
                for lb in asp_obj.levels:
                    if la == lb:
                        continue
                    rel = tmp.get_aspect_level_relation(aspect_name, la, lb)
                    if rel in (BT, BTE, EQ):
                        rel_details.append(f"{la}{rel}{lb}")
            result["imported_aspects"].append({
                "name":          aspect_name,
                "level_count":   len(asp_obj.levels),
                "levels":        list(asp_obj.levels.keys()),
                "has_relations": has_relations,
                "relations":     rel_details,
            })

        if result["aspect_errors"]:
            return result

        # ── Step 1.5: import |VDCM| tab if present ────────────────
        vdcm_sheet = next((n for n in wb.sheetnames if n == VDCM), None)
        if vdcm_sheet is not None:
            vdcm_result = tmp.import_vdiff_comparison_matrix_from_worksheet(
                wb[vdcm_sheet]
            )
            result["vdcm_adds"] = len(vdcm_result["adds"])
            result["vdcm_add_details"] = [
                f"{repr(add[0])} {add[1]} {repr(add[2])}"
                for add in vdcm_result["adds"]
                if len(add) >= 3
            ]
            if vdcm_result["collisions"]:
                for coll in vdcm_result["collisions"]:
                    vd1, old_rel, vd2, new_rel = coll
                    result["closure_collisions"].append(
                        f"|VDCM| collision: attempted {repr(vd1)} {new_rel} {repr(vd2)} "
                        f"conflicts with existing {repr(vd1)} {old_rel} {repr(vd2)}"
                    )
                return result

        # ── Step 2: closure check ───────────────────────────────────
        _, _, closure_colls = tmp.closure()
        if closure_colls:
            for coll_entry in closure_colls:
                origin_type, origin_detail, coll = coll_entry
                vd1, old_rel, vd2, new_rel = coll
                result["closure_collisions"].append(
                    f"Closure collision: "
                    f"attempted {repr(vd1)} {new_rel} {repr(vd2)} "
                    f"conflicts with existing {repr(vd1)} {old_rel} {repr(vd2)}"
                )
            return result

        # ── Step 3: consequences (strict 3a) ────────────────────────
        cons_sheet_name = next((n for n in wb.sheetnames if n == CONS), None)
        if cons_sheet_name is None:
            result["missing_cons_sheet"] = True
        else:
            ws_cons = wb[cons_sheet_name]
            # Read aspect names from header row
            aspect_names = []
            for col_index in range(2, ws_cons.max_column + 1):
                a_name = ws_cons.cell(row=1, column=col_index).value
                if not a_name:
                    break
                aspect_names.append(a_name)
            # Validate and stage consequences
            for row in ws_cons.iter_rows(min_row=3, values_only=True):
                short_name = row[0]
                if not short_name:
                    break
                aspect_levels = {
                    a: str(row[i + 1])
                    for i, a in enumerate(aspect_names)
                }
                # 3a strict: check all levels exist before adding
                unknown = []
                for a_name, level in aspect_levels.items():
                    if not tmp.has_aspect(a_name):
                        unknown.append(f"unknown aspect '{a_name}'")
                    elif level not in tmp.aspects[a_name].levels:
                        unknown.append(f"unknown level '{level}' for aspect '{a_name}'")
                if unknown:
                    result["consequence_errors"].append(
                        f"'{short_name}': " + ", ".join(unknown)
                    )
                else:
                    try:
                        tmp.add_consequence(short_name, aspect_levels)
                        result["imported_consequences"].append(short_name)
                        result["imported_consequence_details"].append({
                            "name": short_name,
                            "repr": str(tmp.consequences[short_name])
                        })
                    except ValueError as e:
                        result["consequence_errors"].append(f"'{short_name}': {e}")
            if result["consequence_errors"]:
                return result

        # ── Step 4: commit ──────────────────────────────────────────
        self.aspects                 = tmp.aspects
        self.consequences            = tmp.consequences
        self.vdiff_comparison_matrix = tmp.vdiff_comparison_matrix
        result["success"] = True
        return result

    def import_aspect_from_worksheet(self, ws) -> List:
        aspect_name = ws["A1"].value
        data_type_str = ws["B1"].value
        description = ws["A2"].value
        logger.info(f"Importing aspect '{aspect_name}'")
        # Create aspect and add levels
        self.add_aspect(aspect_name, data_type_str, description)
        rows = [["Aspect:", aspect_name], ["Type:", data_type_str], ["Description:", description]]
        # Read levels from row 3 onwards
        for row in ws.iter_rows(min_row=3, values_only=True):
            level = row[0]
            if level is None:
                break
            description = row[1] if len(row) > 1 else None
            self.add_aspect_level(aspect_name, level, description)
            rows.append([level, description])
        return rows

    def import_aspect_level_relations_from_worksheet(self, ws) -> Tuple:
        aspect_name = ws["A1"].value
        aspect = self.get_aspect(aspect_name)
        col_heads = []
        for col in ws.iter_cols(min_col=5, values_only=True):
            key_c = col[1]
            col_heads.append(key_c)
        adds, colls = [], []
        for row in ws.iter_rows(min_row=3, values_only=True):
            key_r = row[3]
            col_index = 4
            for key_c in col_heads:
                rel = row[col_index] if row[col_index] else UNDEFINED
                (a, c) = self.set_aspect_level_relation(aspect_name, key_r, key_c, rel)
                adds += a
                colls += c
                col_index += 1
        logger.debug(str(adds) + " " + str(colls))
        return (adds, colls)
    
    def import_consequences_from_worksheet(self, ws):
        aspect_names = []
        for col_index in range(2, ws.max_column + 1):
            aspect_name = ws.cell(row=1, column=col_index).value
            aspect_type_str = ws.cell(row=2, column=col_index).value
            if not aspect_name:
                break
            else:
                if not self.has_aspect(aspect_name):
                    aspect_type_str = aspect_type_str if aspect_type_str else "str"
                    added_aspect = self.add_aspect(aspect_name, aspect_type_str, None)
                    logger.debug(str(added_aspect))
                aspect_names.append(aspect_name)
        for row in ws.iter_rows(min_row=3, values_only=True):
            short_name = row[0]
            if not short_name:
                break
            aspect_levels = {}
            for col_index, aspect_name in enumerate(aspect_names, start=1):
                level_str = str(row[col_index])
                aspect_levels[aspect_name] = level_str
            self.add_consequence(short_name, aspect_levels)

    def vdiff_comparison_matrix_str(self, vdcm):
        result = ""
        for vd1, row in vdcm.items():
            for vd2, rel in row.items():
                result += f"   {vd1} {GTE} {vd2}: {rel}\n"
        return result

    def __repr__(self):
        result = "Aspects:\n"
        for aspect in self.aspects.values():
            result += "- " + str(aspect) + "\n"
        result += "\nConsequence space:\n"
        for consequence in self.consequence_space:
            result += " - " + str(consequence) + "\n"

        result += "\nConsequences:\n"
        for (short_name, c) in self.consequences.items():
            result += " - " + short_name + ":" + str(c) + "\n"
        result += "\nValue-difference relations:\n"
        result += self.vdiff_comparison_matrix_str(self.vdiff_comparison_matrix)
        return result

    def to_dict(self):
        """Full JSON-serialization of the entire EudoxaManager."""

        def _vd_serial(vd: VDiff) -> str:
            """Encode a VDiff as a JSON-safe string key.
            NATURAL_ZERO → ZDIFF_DISPLAY ('◬').
            Other vdiffs → 'aspect_name|||from_level|||to_level'.
            ('|||' is forbidden in aspect names, and level names that contain
            '|||' are extremely unlikely in practice.)
            """
            if vd is NATURAL_ZERO or vd.natural_zero():
                return ZDIFF_DISPLAY
            an = vd.aspect_name or ""
            fl = "" if vd.from_level is None else str(vd.from_level)
            tl = "" if vd.to_level  is None else str(vd.to_level)
            return f"{an}|||{fl}|||{tl}"

        vdcm_out = {}
        for vd1, row in self.vdiff_comparison_matrix.items():
            k1 = _vd_serial(vd1)
            vdcm_out[k1] = {_vd_serial(vd2): rel for vd2, rel in row.items()}

        return {
            "__schema__": 2,
            "aspects": {
                name: aspect.to_dict()
                for name, aspect in self.aspects.items()
            },
            "consequences": {
                short: c.to_dict()
                for short, c in self.consequences.items()
            },
            "vdiff_comparison_matrix": vdcm_out
        }

    @classmethod
    def from_dict(cls, data):
        mgr = cls()

        # Clear auto-generated initial values
        mgr.aspects = {}
        mgr.consequences = {}
        mgr.vdiff_comparison_matrix = {}

        # ---- Aspects ----
        for name, asp_data in data.get("aspects", {}).items():
            mgr.aspects[name] = Aspect.from_dict(asp_data)

        # ---- Consequences ----
        for short, cons_data in data.get("consequences", {}).items():
            mgr.consequences[short] = Consequence.from_dict(cons_data)

        # ---- VDiff comparison matrix ----
        schema = data.get("__schema__", 1)
        vdcm_in = data.get("vdiff_comparison_matrix", {})

        if schema >= 2:
            # Schema 2: keys are '_vd_serial' strings produced by to_dict.
            def _vd_parse(key: str) -> VDiff:
                if key == ZDIFF_DISPLAY:
                    return NATURAL_ZERO
                parts = key.split("|||")
                an = parts[0] or None
                fl = parts[1] or None
                tl = parts[2] or None
                return VDiff(an, fl, tl)

            for k1, row in vdcm_in.items():
                vd1 = _vd_parse(k1)
                mgr.vdiff_comparison_matrix[vd1] = {
                    _vd_parse(k2): rel for k2, rel in row.items()
                }

        else:
            # Schema 1 (legacy): outer keys are "a1|||a2", inner keys are
            # "f1::t1>>f2::t2".  Migrate by normalising natural zeros to
            # NATURAL_ZERO; duplicate entries (e.g. the same NATURAL_ZERO pair
            # appearing under several aspect-pair outer keys) are idempotent.
            vdcm = mgr.vdiff_comparison_matrix
            for key, relation_map in vdcm_in.items():
                a1, a2 = key.split("|||")
                for pair_key, rel in relation_map.items():
                    d1s, d2s = pair_key.split(">>")
                    f1, t1 = d1s.split("::")
                    f2, t2 = d2s.split("::")
                    f1 = None if f1 == "" else f1
                    t1 = None if t1 == "" else t1
                    f2 = None if f2 == "" else f2
                    t2 = None if t2 == "" else t2
                    k1 = _vdiff_key(VDiff(a1, f1, t1))
                    k2 = _vdiff_key(VDiff(a2, f2, t2))
                    if k1 not in vdcm:
                        vdcm[k1] = {}
                    vdcm[k1][k2] = rel

        return mgr